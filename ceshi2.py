from openpyxl import load_workbook

excel_path = r"D:\Desktop\宁波银行.xlsx"
wb = load_workbook(excel_path)
sheet = wb["宁波银行"]

# 从第2行开始读取（跳过表头）
for row in sheet.iter_rows(min_row=2, values_only=True):
    xiangmu, yinhangkahao = row
    if not xiangmu or not yinhangkahao:
        continue  # 跳过空行

    print(f"项目名称：{xiangmu}")
    print(f"银行账号：{yinhangkahao}")

    # 这里可以插入你的自动化下载逻辑
    # run_automation(xiangmu, yinhangkahao)
